"""
Builder Module
Converts normalized pricing records and add-ons into a formatted Excel workbook.

Sheets produced:
  - Master Pricing        (all records)
  - Group A Slabs
  - Group A Bifolds
  - Group B Slabs
  - Group B Bifolds
  - Group K Slabs
  - Group K Bifolds
  - Trimlite Slabs
  - PQ East Slabs
  - Add-ons - All Groups

Each sheet is formatted with a header row and alternating row colours.
"""

import os
from typing import List, Dict, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ──────────────────────────────────────────────────────────────────────────────
# Colour palette
# ──────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark blue
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6") # mid blue
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")       # light blue
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="BBBBBB")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ──────────────────────────────────────────────────────────────────────────────
# DataFrame construction helpers
# ──────────────────────────────────────────────────────────────────────────────

RECORD_COLUMNS = [
    "dealer_group",
    "product_type",
    "style",
    "size",
    "variant",
    "price",
    "price_numeric",
    "source_pdf",
]

ADDON_COLUMNS = [
    "dealer_group",
    "addon_name",
    "price",
    "price_numeric",
    "description",
    "applicable_to",
    "source_pdf",
]


def _records_to_df(records: List[Dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=RECORD_COLUMNS)
    df = pd.DataFrame(records)
    for col in RECORD_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    # Serialize the 'options' list if present
    if "options" in df.columns:
        df["options"] = df["options"].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
    return df[RECORD_COLUMNS]


def _addons_to_df(addons: List[Dict]) -> pd.DataFrame:
    if not addons:
        return pd.DataFrame(columns=ADDON_COLUMNS)
    df = pd.DataFrame(addons)
    for col in ADDON_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    if "applicable_to" in df.columns:
        df["applicable_to"] = df["applicable_to"].apply(
            lambda x: ", ".join(x) if isinstance(x, list) else x
        )
    return df[ADDON_COLUMNS]


# ──────────────────────────────────────────────────────────────────────────────
# Excel formatting helpers
# ──────────────────────────────────────────────────────────────────────────────

def _write_df_to_sheet(ws, df: pd.DataFrame, title: str) -> None:
    """Write a DataFrame to an openpyxl worksheet with formatting."""
    # Title row
    ws.append([title])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="left")
    ws.append([])  # blank

    if df.empty:
        ws.append(["No data extracted for this section."])
        return

    # Header row
    header_row_num = ws.max_row + 1
    headers = list(df.columns)
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_num, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_offset, row_data in enumerate(df.itertuples(index=False), start=1):
        row_num = header_row_num + row_offset
        fill = ALT_FILL if row_offset % 2 == 0 else WHITE_FILL
        ws.append(list(row_data))
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = THIN_BORDER

    # Auto-width columns
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in ws.iter_rows(
            min_row=header_row_num, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Freeze panes below header
    ws.freeze_panes = ws.cell(row=header_row_num + 1, column=1)


# ──────────────────────────────────────────────────────────────────────────────
# Main builder
# ──────────────────────────────────────────────────────────────────────────────

def build_excel(
    data: Dict[str, List],
    output_path: str,
) -> str:
    """
    Build a formatted Excel workbook from parsed pricing data.

    Args:
        data: {'records': [...], 'addons': [...]}
        output_path: File path for the output .xlsx

    Returns:
        Absolute path of the written file.
    """
    records = data.get("records", [])
    addons = data.get("addons", [])

    all_df = _records_to_df(records)
    addons_df = _addons_to_df(addons)

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Master Pricing ──────────────────────────────────────────────────────
    ws_master = wb.create_sheet("Master Pricing")
    _write_df_to_sheet(ws_master, all_df, "Master Pricing Table – All Groups")

    # ── Per dealer group / product type sheets ───────────────────────────────
    groups = [
        ("Group A", "door",   "Group A Slabs"),
        ("Group A", "bifold", "Group A Bifolds"),
        ("Group B", "door",   "Group B Slabs"),
        ("Group B", "bifold", "Group B Bifolds"),
        ("Group K", "door",   "Group K Slabs"),
        ("Group K", "bifold", "Group K Bifolds"),
        ("Trimlite", "door",  "Trimlite Slabs"),
        ("PQ East", "door",   "PQ East Slabs"),
    ]

    for dealer, ptype, sheet_name in groups:
        sub = all_df[
            (all_df["dealer_group"] == dealer) & (all_df["product_type"] == ptype)
        ].copy()
        ws = wb.create_sheet(sheet_name[:31])
        _write_df_to_sheet(ws, sub, f"{sheet_name}")

    # ── Add-ons sheet ────────────────────────────────────────────────────────
    ws_addons = wb.create_sheet("Add-ons - All Groups")
    _write_df_to_sheet(ws_addons, addons_df, "Add-ons / Machining / Jambs / Options")

    # ── Per-group add-on sheets ──────────────────────────────────────────────
    for dealer in ["Group A", "Group B", "Group K", "Trimlite", "PQ East"]:
        sub = addons_df[addons_df["dealer_group"] == dealer].copy()
        if not sub.empty:
            safe_name = f"Add-ons {dealer}"[:31]
            ws = wb.create_sheet(safe_name)
            _write_df_to_sheet(ws, sub, f"Add-ons – {dealer}")

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary", 0)
    summary_rows = [
        ["JELD-WEN Interior Door Quoting – Pricing Workbook"],
        [""],
        ["Sheet", "Record Count", "Description"],
        ["Master Pricing", len(all_df), "All pricing records from all PDFs"],
    ]
    for dealer, ptype, sheet_name in groups:
        n = len(
            all_df[
                (all_df["dealer_group"] == dealer) & (all_df["product_type"] == ptype)
            ]
        )
        summary_rows.append([sheet_name, n, f"{dealer} – {ptype} prices"])
    summary_rows.append(["Add-ons - All Groups", len(addons_df), "All add-ons/machining/jambs"])
    summary_rows.append([""])
    summary_rows.append(["Source PDFs processed:", len(set(all_df["source_pdf"].tolist() + addons_df["source_pdf"].tolist() if not all_df.empty else [])), ""])

    for r in summary_rows:
        ws_summary.append(r)

    # Format summary
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws_summary["A3"].font = HEADER_FONT
    ws_summary["B3"].font = HEADER_FONT
    ws_summary["C3"].font = HEADER_FONT
    ws_summary["A3"].fill = HEADER_FILL
    ws_summary["B3"].fill = HEADER_FILL
    ws_summary["C3"].fill = HEADER_FILL
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 45

    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\n  [OK] Excel saved to: {output_path}")
    return output_path
