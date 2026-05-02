"""
Quote Engine Module
Loads extracted JSON data and provides price lookup + quote calculation.

Usage:
    engine = QuoteEngine("output/pricing_data.json")
    groups  = engine.get_dealer_groups()
    styles  = engine.get_styles("Group A")
    sizes   = engine.get_sizes("Group A", "COLONIST TEXT")
    addons  = engine.get_addons("Group A")
    quote   = engine.build_quote(
                  dealer_group="Group A",
                  product_type="door",
                  style="COLONIST TEXT",
                  size='12"',
                  variant="HC",
                  selected_addons=["Machining for 1 3/8\": 3 Hinges and Lock"],
                  quantity=10,
              )
"""

import json
import os
from typing import Any, Dict, List, Optional


class QuoteEngine:
    """Pricing lookup and quote calculation engine."""

    def __init__(self, data_path: str):
        """
        Load pricing data from JSON file produced by the extraction pipeline.

        Args:
            data_path: Path to pricing_data.json
        """
        if not os.path.isfile(data_path):
            raise FileNotFoundError(
                f"Pricing data not found at '{data_path}'. "
                "Run the extraction pipeline first."
            )

        with open(data_path, "r", encoding="utf-8") as fh:
            raw = json.load(fh)

        self._records: List[Dict] = raw.get("records", [])
        self._addons: List[Dict] = raw.get("addons", [])

        # Build lookup index  {dealer_group -> {product_type -> {style -> {size -> {variant -> record}}}}}
        self._index: Dict = {}
        for rec in self._records:
            dg = rec.get("dealer_group", "")
            pt = rec.get("product_type", "")
            st = rec.get("style", "")
            sz = rec.get("size", "")
            va = rec.get("variant", "")
            (
                self._index
                .setdefault(dg, {})
                .setdefault(pt, {})
                .setdefault(st, {})
                .setdefault(sz, {})[va]
            ) = rec

        # Addon index  {dealer_group -> [addon, ...]}
        self._addon_index: Dict[str, List[Dict]] = {}
        for addon in self._addons:
            dg = addon.get("dealer_group", "")
            self._addon_index.setdefault(dg, []).append(addon)

    # ── Discovery methods ──────────────────────────────────────────────────────

    def get_dealer_groups(self) -> List[str]:
        """Return sorted list of all dealer groups."""
        return sorted(self._index.keys())

    def get_product_types(self, dealer_group: str) -> List[str]:
        """Return product types available for a dealer group."""
        return sorted(self._index.get(dealer_group, {}).keys())

    def get_styles(self, dealer_group: str, product_type: str = "door") -> List[str]:
        """Return sorted list of styles for a dealer group and product type."""
        return sorted(
            self._index.get(dealer_group, {}).get(product_type, {}).keys()
        )

    def get_sizes(
        self, dealer_group: str, style: str, product_type: str = "door"
    ) -> List[str]:
        """Return list of sizes for a dealer group, style, and product type."""
        return list(
            self._index.get(dealer_group, {})
            .get(product_type, {})
            .get(style, {})
            .keys()
        )

    def get_variants(
        self, dealer_group: str, style: str, size: str, product_type: str = "door"
    ) -> List[str]:
        """Return available variants for a specific product."""
        return list(
            self._index.get(dealer_group, {})
            .get(product_type, {})
            .get(style, {})
            .get(size, {})
            .keys()
        )

    def get_addons(self, dealer_group: str) -> List[Dict]:
        """Return all add-ons available for a dealer group."""
        return self._addon_index.get(dealer_group, [])

    # ── Price lookup ───────────────────────────────────────────────────────────

    def get_price(
        self,
        dealer_group: str,
        style: str,
        size: str,
        variant: str = "HC",
        product_type: str = "door",
    ) -> Optional[Dict]:
        """
        Look up a specific price record.

        Returns the record dict or None if not found.
        """
        return (
            self._index.get(dealer_group, {})
            .get(product_type, {})
            .get(style, {})
            .get(size, {})
            .get(variant)
        )

    def find_price_fuzzy(
        self,
        dealer_group: str,
        style: str,
        size: str,
        product_type: str = "door",
    ) -> Optional[Dict]:
        """
        Find a price record using case-insensitive matching.
        Returns first matching record (prefers HC variant).
        """
        dg_data = self._index.get(dealer_group, {})
        pt_data = dg_data.get(product_type, {})

        # Case-insensitive style match
        matched_style = None
        for s in pt_data:
            if s.upper() == style.upper():
                matched_style = s
                break
        if matched_style is None:
            return None

        size_data = pt_data[matched_style]

        # Case-insensitive size match
        matched_size = None
        for sz in size_data:
            if sz.upper() == size.upper():
                matched_size = sz
                break
        if matched_size is None:
            return None

        variants = size_data[matched_size]

        # Prefer "HC" variant, then "Full TL", then first available
        for preferred in ("HC", "HC c/w Hardware", "Full TL", "LTL"):
            if preferred in variants:
                return variants[preferred]
        return next(iter(variants.values()), None)

    # ── Quote builder ──────────────────────────────────────────────────────────

    def build_quote(
        self,
        dealer_group: str,
        product_type: str,
        style: str,
        size: str,
        variant: str = "HC",
        selected_addons: Optional[List[str]] = None,
        quantity: int = 1,
    ) -> Dict[str, Any]:
        """
        Build a complete quote for a door/bifold with optional add-ons.

        Returns a quote dict with itemised breakdown and totals.
        """
        quote: Dict[str, Any] = {
            "dealer_group": dealer_group,
            "product_type": product_type,
            "style": style,
            "size": size,
            "variant": variant,
            "quantity": quantity,
            "line_items": [],
            "subtotal": 0.0,
            "total": 0.0,
            "errors": [],
        }

        # Base price
        rec = self.get_price(dealer_group, style, size, variant, product_type)
        if rec is None:
            rec = self.find_price_fuzzy(dealer_group, style, size, product_type)

        if rec is None:
            quote["errors"].append(
                f"NOT FOUND IN PDF: No price for {dealer_group} / {style} / {size} / {variant}"
            )
            base_price = 0.0
        else:
            pn = rec.get("price_numeric")
            if pn is None:
                quote["errors"].append(
                    f"Price not numeric for {dealer_group} / {style} / {size}: {rec.get('price')}"
                )
                base_price = 0.0
            else:
                base_price = pn

        quote["line_items"].append(
            {
                "description": f"{style} {size} ({variant})",
                "unit_price": base_price,
                "quantity": quantity,
                "subtotal": round(base_price * quantity, 2),
            }
        )
        running_total = base_price * quantity

        # Add-ons
        if selected_addons:
            available = {a["addon_name"]: a for a in self.get_addons(dealer_group)}
            for addon_name in selected_addons:
                addon = available.get(addon_name)
                if addon is None:
                    quote["errors"].append(f"Add-on not found: '{addon_name}'")
                    continue
                apn = addon.get("price_numeric")
                if apn is None:
                    quote["errors"].append(
                        f"Add-on price not numeric: '{addon_name}' → {addon.get('price')}"
                    )
                    continue
                addon_total = round(apn * quantity, 2)
                quote["line_items"].append(
                    {
                        "description": f"Add-on: {addon_name}",
                        "unit_price": apn,
                        "quantity": quantity,
                        "subtotal": addon_total,
                    }
                )
                running_total += addon_total

        quote["subtotal"] = round(running_total, 2)
        quote["total"] = quote["subtotal"]
        return quote

    def export_quote_to_excel(self, quote: Dict[str, Any], output_path: str) -> str:
        """
        Export a single quote to an Excel file.

        Returns the path of the saved file.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Quote"

        DARK = PatternFill("solid", fgColor="1F4E79")
        LIGHT = PatternFill("solid", fgColor="D6E4F0")
        BOLD_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        BOLD_DARK = Font(name="Calibri", bold=True, color="1F4E79", size=11)
        NORMAL = Font(name="Calibri", size=10)

        # Title
        ws["A1"] = "JELD-WEN Interior Door Quote"
        ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
        ws.append([])

        # Meta info
        meta = [
            ("Dealer Group", quote["dealer_group"]),
            ("Product Type", quote["product_type"]),
            ("Style", quote["style"]),
            ("Size", quote["size"]),
            ("Variant", quote["variant"]),
            ("Quantity", quote["quantity"]),
        ]
        for label, val in meta:
            ws.append([label, val])
            ws.cell(row=ws.max_row, column=1).font = Font(name="Calibri", bold=True, size=10)

        ws.append([])

        # Line items header
        hdr_row = ws.max_row + 1
        ws.append(["Description", "Unit Price", "Quantity", "Subtotal"])
        for col in range(1, 5):
            c = ws.cell(row=hdr_row, column=col)
            c.font = BOLD_WHITE
            c.fill = DARK
            c.alignment = Alignment(horizontal="center")

        # Line items
        for i, item in enumerate(quote["line_items"]):
            ws.append([
                item["description"],
                f"${item['unit_price']:.2f}",
                item["quantity"],
                f"${item['subtotal']:.2f}",
            ])
            fill = LIGHT if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for col in range(1, 5):
                c = ws.cell(row=ws.max_row, column=col)
                c.font = NORMAL
                c.fill = fill

        # Total row
        ws.append([])
        total_row = ws.max_row + 1
        ws.append(["", "", "TOTAL", f"${quote['total']:.2f}"])
        for col in (3, 4):
            c = ws.cell(row=total_row, column=col)
            c.font = BOLD_DARK
            c.fill = PatternFill("solid", fgColor="BDD7EE")

        # Errors
        if quote.get("errors"):
            ws.append([])
            ws.append(["NOTES / WARNINGS"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="C00000")
            for err in quote["errors"]:
                ws.append(["", err])

        # Column widths
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 14

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        return output_path
